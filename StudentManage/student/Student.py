#  __*_coding:utf-8_*_
# Time :        2020/8/5 16:29
# Author:       LeiLei
# File:         Student.py
# Software:     PyCharm
from PyQt5.QtCore import QModelIndex
from PyQt5.QtWidgets import QWidget, QMessageBox, QHeaderView, QPushButton, QHBoxLayout, QTableWidgetItem, QFileDialog
from PyQt5.QtGui import QStandardItem, QStandardItemModel, QColor
from openpyxl import Workbook, load_workbook
import mysql.connector as mc
from StudentManage.frame.student import Ui_Form
from StudentManage.config import *


def get_row(sheet, row):
    col = sheet.max_column
    result = []
    for i in range(1, col+1):
        result.append(sheet.cell(row=row, column=i).value)
    return result


class Student(QWidget, Ui_Form):
    def __init__(self):
        super(Student, self).__init__()
        self.setupUi(self)
        self.model_1 = QStandardItemModel()
        self.model_1.setHorizontalHeaderLabels(TABLE_TITLE)
        self.tableView.setModel(self.model_1)
        self.tableWidget_2.setColumnCount(len(TABLE_TITLE))
        self.tableWidget_3.setColumnCount(len(TABLE_TITLE))
        self.tableWidget_2.setHorizontalHeaderLabels(TABLE_TITLE)
        self.tableWidget_3.setHorizontalHeaderLabels(TABLE_TITLE)
        self.tableWidget.setColumnCount(len(TABLE_TITLE_2))
        self.tableWidget.setHorizontalHeaderLabels(TABLE_TITLE_2)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_3.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.header_v = self.tableView.verticalHeader()
        self.header_h = self.tableView.horizontalHeader()
        self.header_h.setSectionResizeMode(QHeaderView.Stretch)
        self.header_v.hide()
        self.conn = mc.connect(host='localhost', user='Lei', password='1314520', db='study')
        self.cursor = self.conn.cursor()
        self.sexName = '男'
        if not self.conn.is_connected():
            QMessageBox.warning(self, "警告", "无法连接到服务器", QMessageBox.Ok)
        self.lineEdit_6.setText("信息工程学院")

    def sex(self):
        if self.radioButton_2.isChecked() or self.radioButton_4.isChecked():
            self.sexName = '男'
        elif self.radioButton.isChecked() or self.radioButton_3.isChecked():
            self.sexName = '女'

    def add(self):
        if self.lineEdit_3.text() == '':
            QMessageBox.warning(self, '提示', '学号不能为空', QMessageBox.Ok)
            return
        if self.lineEdit_4.text() == '':
            QMessageBox.warning(self, '提示', '姓名不能为空', QMessageBox.Ok)
            return
        if self.lineEdit_5.text() == '':
            QMessageBox.warning(self, '提示', '出生日期不能为空', QMessageBox.Ok)
            return
        if self.lineEdit_7.text() == '':
            QMessageBox.warning(self, '提示', '电话号码不能为空', QMessageBox.Ok)
            return
        if self.lineEdit_8.text() == '':
            QMessageBox.warning(self, '提示', '学院名称不能为空', QMessageBox.Ok)
            return
        command = "insert into student(stu_id, stu_name, stu_sex, stu_birth, stu_tel, stu_college)" \
                  "values ('%s', '%s', '%s', '%s', '%s', '%s')" % \
                  (self.lineEdit_3.text(), self.lineEdit_4.text(), self.sexName, self.lineEdit_5.text(), self.lineEdit_7.text(), self.lineEdit_8.text())
        try:
            self.cursor.execute(command)
            self.conn.commit()
            QMessageBox.information(self, '提示', '添加成功', QMessageBox.Ok)
        except Exception as e:
            QMessageBox.warning(self, '提示', '添加错误' + e.args[1], QMessageBox.Ok)

    def lead_in(self):
        pass

    def lead_out(self):
        path = QFileDialog.getSaveFileName(self, 'save file', '', 'Excel File(*.xlsx)')
        book = Workbook()
        sheet = book.active
        for i in range(self.tableWidget_3.rowCount()+1):
            for j in range(self.tableWidget_3.columnCount()):
                if i == 0:
                    sheet.cell(row=i+1, column=j+1, value=TABLE_TITLE[j])
                else:
                    sheet.cell(row=i + 1, column=j + 1, value=self.tableWidget_3.item(i-1, j).text())
        book.save(path[0])
        book.close()
        QMessageBox.information(self, '提示', '已经成功导出', QMessageBox.Ok)

    def modify(self):
        stu_id = self.tableWidget_2.item(0, 0).text()
        stu_sex = self.sexName
        command = "update student set stu_sex='%s', stu_birth='%s', stu_tel='%s', stu_college='%s' where stu_id='%s'" \
                  % (stu_sex, self.lineEdit_11.text(), self.lineEdit_12.text(), self.lineEdit_13.text(), stu_id)
        self.cursor.execute(command)
        self.conn.commit()
        QMessageBox.information(self, '提示', '数据修改成功', QMessageBox.Ok)

    def choose_file(self):
        path = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel File(*.xlsx)')
        self.lineEdit_9.setText(path[0])
        book = load_workbook(path[0])
        sheet = book.active
        if get_row(sheet, 1) != TABLE_TITLE:
            QMessageBox.warning(self, '提示', '请按照标准的格式调整Excel文件', QMessageBox.Ok)
            return
        command = "insert into student(stu_id, stu_name, stu_sex, stu_birth, stu_tel, stu_college) values ('%s', '%s', '%s', '%s', '%s', '%s')"
        for i in range(2, sheet.max_row+1):
            rows = get_row(sheet, i)
            com = command % tuple(rows)
            try:
                self.cursor.execute(com)
                self.conn.commit()
                QMessageBox.information(self, '提示', '数据插入成功', QMessageBox.Ok)
            except Exception:
                pass

    def delete(self):
        row = int(self.sender().text().replace('删除', ''))
        command = "delete from student where stu_id='%s'" % self.tableWidget.item(row, 0)
        self.cursor.execute(command)
        self.conn.commit()
        self.tableWidget.removeRow(row)

    def modify_search(self):
        self.tableWidget_2.clearContents()
        text = self.lineEdit_10.text()
        self.lineEdit_10.setText('')
        if text == '':
            return
        result = self.base_search(text, False)
        self.tableWidget_2.setRowCount(len(result))
        for i in range(self.tableWidget_2.rowCount()):
            self.tableWidget_2.setRowHeight(i, 60)
            for j in range(self.tableWidget_2.columnCount()):
                if j == 2:
                    if result[i][j] == '男':
                        self.radioButton_4.setChecked(True)
                    else:
                        self.radioButton_3.setChecked(True)
                elif j == 3:
                    self.lineEdit_11.setText(result[i][j].strftime('%Y-%m-%d'))
                elif j == 4:
                    self.lineEdit_12.setText(result[i][j])
                elif j == 5:
                    self.lineEdit_13.setText(result[i][j])
                if type(result[i][j]) is not str:
                    item = QTableWidgetItem(result[i][j].strftime('%Y-%m-%d'))
                else:
                    item = QTableWidgetItem(result[i][j])
                item.setBackground(QColor('green'))
                self.tableWidget_2.setItem(i, j, item)

    def delete_search(self):
        self.tableWidget.clearContents()
        text = self.lineEdit_2.text()
        self.lineEdit_2.setText('')
        if text == '':
            return
        result = self.base_search(text)
        self.tableWidget.setRowCount(len(result))
        for i in range(self.tableWidget.rowCount()):
            self.tableWidget.setRowHeight(i, 60)
            for j in range(self.tableWidget.columnCount()):
                if j == self.tableWidget.columnCount()-1:
                    bnt = QPushButton("删除%d" % i)
                    bnt.setStyleSheet(''' text-align : center;background-color : blue;height : 50px;font : 16px; color: white;''')
                    bnt.clicked.connect(self.delete)
                    layout_h = QHBoxLayout()
                    layout_h.addWidget(bnt)
                    widget = QWidget()
                    widget.setLayout(layout_h)
                    self.tableWidget.setCellWidget(i, j, widget)
                    break
                if type(result[i][j]) is not str:
                    item = QTableWidgetItem(result[i][j].strftime('%Y-%m-%d'))
                else:
                    item = QTableWidgetItem(result[i][j])
                item.setBackground(QColor('green'))
                self.tableWidget.setItem(i, j, item)

    def output_search(self):
        self.tableWidget_3.clearContents()
        text = self.lineEdit_6.text()
        self.lineEdit_6.setText('')
        if text == '':
            return
        result = self.base_search(text)
        self.tableWidget_3.setRowCount(len(result))
        for i in range(self.tableWidget_3.rowCount()):
            self.tableWidget_3.setRowHeight(i, 60)
            for j in range(self.tableWidget_3.columnCount()):
                if type(result[i][j]) is not str:
                    item = QTableWidgetItem(result[i][j].strftime('%Y-%m-%d'))
                else:
                    item = QTableWidgetItem(result[i][j])
                item.setBackground(QColor('green'))
                self.tableWidget_3.setItem(i, j, item)

    def base_search(self, text, flag=True):
        if flag:
            command = "select * from student where stu_id='%s' or stu_college='%s' or stu_name='%s'" % (text, text, text)
        else:
            command = "select * from student where stu_id='%s'" % text
        self.cursor.execute(command)
        result = self.cursor.fetchall()
        return result

    def search(self):
        self.model_1.clear()
        self.model_1.setHorizontalHeaderLabels(TABLE_TITLE)
        text = self.lineEdit.text()
        self.lineEdit.setText('')
        if text == '':
            return
        result = self.base_search(text)
        for i in range(len(result)):
            for j in range(len(result[i])):
                if type(result[i][j]) is not str:
                    item = QStandardItem(result[i][j].strftime('%Y-%m-%d'))
                else:
                    item = QStandardItem(result[i][j])
                item.setBackground(QColor('green'))
                self.model_1.setItem(i, j, item)
